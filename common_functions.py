import mysql.connector
from openpyxl import load_workbook
import random, urllib, json, string, csv
from SPARQLWrapper import SPARQLWrapper, JSON
import time
import numpy as np
import requests
import re
import sys

from pdb import set_trace as st

def connect_to_sql(db="grounds_sshoc"):
    mydb = mysql.connector.connect(
        host="round4",
        user="sshoc",
        password="IqazZuKaXMmSEqUl",
        database=db
    )

    return mydb

def get_json(url):
    operUrl = urllib.request.urlopen(url)
    if(operUrl.getcode()==200):
       data = operUrl.read()
       json_data = json.loads(data)
    else:
       print("Error receiving data", operUrl.getcode())
    return json_data
    
def find_gallery_PID(ng_number):
    ng_number = str(ng_number)
    if ng_number.startswith('https'):
        ng_number = str(ng_number.rsplit('/',1)[-1])
    else:
        ng_number = ng_number.replace(' ','_')

    gallery_PID = None
    
    # Added as the NG code needs to be formatted correctly to match the ES system
    code = UT_split_code(ng_number);
    #print (ng_number + ' ----> ' + code[14])
    
    if code[14] != "NG0":
      ng_number = code[14]

    try:
        export_url = 'https://collectiondata.ng-london.org.uk/es/ng-public/_search?q=identifier.object_number:' + ng_number
    except:
        export_url is None
        
    if export_url is not None:
        data = requests.get(export_url, verify=False)
        json = data.json()
        if 'error' not in json:
            if json['hits']['total'] > 0:
                try:
                    json_ng_number = json['hits']['hits'][0]['_source']['identifier'][0]['object_number']
                except:
                    json_ng_number = None
                if json_ng_number == ng_number:
                    gallery_PID = json['hits']['hits'][0]['_source']['identifier'][1]['pid_tms']
    
    return gallery_PID

def check_db(input_literal, table_name):
    db = connect_to_sql()
    cursor = db.cursor()
    pid_value = None
    input_literal = str(input_literal)
    if table_name == 'temp_pids':
        val = 'pid_value'
        str_input = 'literal_value'
    elif table_name == 'wikidata':
        val = 'wd_value'
        str_input = 'string_literal'

    query = "SELECT " + val + " FROM sshoc." + table_name + " WHERE " + str_input + " = %s"
    cursor.execute(query, (input_literal,))
    result = cursor.fetchall()
    
    if len(result) > 0:
        pid_value = result[0][0]

    return pid_value

def generate_placeholder_PID(input_literal):
    N = 4
    placeholder_PID = ""
    for number in range(N):
        res = ''.join(random.choices(string.ascii_uppercase + string.digits, k = N))
        placeholder_PID += str(res)
        placeholder_PID += '-'
    placeholder_PID = placeholder_PID[:-1]
    
    # Added to make them a different structure to real ones.
    placeholder_PID = "SC-" + placeholder_PID
    
    #input_list = [input_literal, placeholder_PID]
    #fields = ['Literal value','Placeholder PID']
    existing_pid = check_db(input_literal, table_name = 'temp_pids')
    old_pid = find_gallery_PID(input_literal)
    
    if existing_pid is not None:
        return existing_pid
    elif old_pid is not None:
        return old_pid
    else:
        db = connect_to_sql()
        cursor = db.cursor()
        input_literal = str(input_literal)
        query = 'INSERT INTO sshoc.temp_pids (literal_value, pid_value) VALUES (%s, %s)'
        val = (input_literal, placeholder_PID)
        cursor.execute(query, val)
        db.commit()

        placeholder_PID = str(placeholder_PID)
        return placeholder_PID

def create_PID_from_triple(pid_type, subj):
    if pid_type == 'object':
        pid_name = subj
    else:
        pid_name = pid_type + ' of ' + subj
    output_pid = generate_placeholder_PID(pid_name)

    return output_pid

def triples_to_csv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.csv','w',newline='') as f:
        write = csv.writer(f)

        write.writerow(fields)
        write.writerows(triples)

    print('CSV created!')
    
def triples_to_tsv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.tsv','w',newline='') as f:
        write = csv.writer(f, delimiter = '\t')

        write.writerow(fields)
        write.writerows(triples)

    print('TSV created!')

def get_property(uri):
    remove_uri = uri.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    final_property = remove_uri.replace('_',' ')
    if '.' in final_property:
        final_property = str(final_property.split('.')[1])
    if 'RRR' in final_property:
        final_property = final_property.replace('RRR','')
    return final_property

def find_aat_value(material,material_type):
    if 'https://rdf.ng-london.org.uk/raphael/resource/' in material:
        material = str(get_property(material))
    else:
        material = str(material)
    
    if material_type == 'medium':
        wb = load_workbook(filename = 'inputs/NG_Meduim_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Medium Material']
    elif material_type == 'support':
        wb = load_workbook(filename = 'inputs/NG_Meduim_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Support Materials']
    elif material_type == 'roles':
        wb = load_workbook(filename = 'inputs/NG_Roles_AAT.xlsx', read_only=True)
        ws = wb['AAT_Roles']
    elif material_type == 'material_grounds':
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['GROUNDS materials']
    elif material_type == 'protocols':
        wb = load_workbook(filename = 'inputs/Protocol_AAT.xlsx', read_only=True)
        ws = wb['Protocol']
    elif material_type == 'techniques':
        wb = load_workbook(filename = 'inputs/Protocol_AAT.xlsx', read_only=True)
        ws = wb['Techniques']
    for row in ws.iter_rows(values_only=True):
        if material in row:
            aat_dict = {}
            aat_number_string = str(row[1])
            aat_number = aat_number_string.replace('aat:','')
            aat_type = row[2]
            aat_dict.update({aat_number:aat_type})
            for x in range(3, 9, 2):
                try:
                    aat_number_string = str(row[x])
                    aat_number = aat_number_string.replace('aat:','')
                    aat_type = row[x+1]
                    aat_dict.update({aat_number:aat_type})
                except:
                    pass
            return aat_dict
    wb.close()

def run_ruby_program(input_string):
    import subprocess

    ruby_var = 'ruby citation_parser.rb \'' + input_string + '\''
    output = subprocess.Popen(ruby_var, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out, error = output.communicate()

    try:
        string_output = out.decode("utf-8")
        json_output = json.loads(string_output)
    except:
        return

    return json_output

def wikidata_query(literal, literal_type, db):
    sparql = SPARQLWrapper('https://query.wikidata.org/sparql')
    print('Trying a query with input ' + literal)
    remove_uri = literal.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    literal = remove_uri.replace('_',' ')
    literal = literal.replace('%C3%A9', 'e')
    literal = literal.replace('%C3%A0', 'a')
    
    if literal_type == 'year':
        thing_type = 'Q577'
        if 'RRR' in literal:
            string_literal = str(literal.rsplit(' ')[-1])
        elif 'About' in literal:
            string_literal = str(literal.rsplit('-')[-1])
        else:
            return None
    elif literal_type == 'institution':
        thing_type = 'Q207694'
        string_literal = str(literal)
    elif literal_type == 'location':
        thing_type = 'Q515'
        string_literal = str(literal)
    elif literal_type == 'language':
        thing_type = 'Q34770'
        string_literal = str(literal)
    
    result = check_db(string_literal, 'wikidata')
    
    if result is not None:
        result = result.replace('http://www.wikidata.org/entity/','')
        return result
    else:
        query = ('''
        SELECT DISTINCT ?year
        WHERE
        {
        ?year  wdt:P31 wd:''' + thing_type + ''' .
                ?year rdfs:label ?yearLabel .
        FILTER( str(?yearLabel) = "''' + string_literal + '''" ) .
        SERVICE wikibase:label { bd:serviceParam wikibase:language "[AUTO_LANGUAGE],en". }
        }
        ''')
        
        sparql.setQuery(query)
        sparql.setReturnFormat(JSON)

        ret = sparql.query().convert()
        time.sleep(3)
        
        if ret["results"]["bindings"] != []:
            for result in ret["results"]["bindings"]:
                #db = connect_to_sql()
                cursor = db.cursor()
                result = result["year"]["value"]
                query = 'INSERT INTO sshoc.wikidata (string_literal, wd_value) VALUES (%s, %s)'
                val = (string_literal, result)
                cursor.execute(query, val)
                db.commit()
        else:
            #db = connect_to_sql()
            cursor = db.cursor()
            query = 'INSERT INTO sshoc.wikidata (string_literal, wd_value) VALUES (%s, %s)'
            val = (string_literal, 'No WD value')
            cursor.execute(query, val)
            db.commit()
        result = result.replace('http://www.wikidata.org/entity/','')
        return result

def create_year_dates(year):
    import datetime
    year = int(year)

    start = datetime.datetime(year, 1, 1)
    end = datetime.datetime(year, 12, 31)

    return start, end

def check_aat_values(col_name):
    aat_db = connect_to_sql(db="sshoc")
    cursor = aat_db.cursor()
    query = "SELECT aat_title, aat_value, aat_unit_title, aat_unit_value FROM aat_refs_grounds WHERE aat_refs_grounds.dimension_column_ref = '" + col_name + "'"
    cursor.execute(query)
    result = cursor.fetchall()

    aat_title, aat_value, aat_unit_title, aat_unit_value = [x for x in result[0]]

    return aat_title, aat_value, aat_unit_title, aat_unit_value

def process_name_prefixes(row):

    if row["person_prefix_name"] is not np.nan:
        value = str(row["person_prefix_name"]) + " " + row["person_name"]
    else:
        value = row["person_name"]

    return value
    
def UT_split_code (ob_code):
	ob_code = ob_code.strip("'" )
	ob_code = ob_code.strip();

	se = re.search('^([a-zA-Z]*)[._-]*([0-9]+)[._-]*([0-9]*)[._-]*([a-zA-Z]+)*[._-]*([0-9]*)(.*)$', ob_code)
	if se:
		var = [se.group(),se.group(1),se.group(2),se.group(3),se.group(4),se.group(5)]
	else:
		var = [False,False,False,False,False,False]
	
	code = [False] * 16
	codeX = ""
	
	if not var[3]:
		var[3] = 0

	er = ["new", "XS", "S", "a", "b", "s"]	
	if var[1] == "NGL":
		code[0] = "%03.0f" % (float(var[2]),)
		code[1] = "= "+str(int(var[3]))
		code[2] = "%02.0f" % (float(var[3]),)
		code[3] = var[1]
	elif var[1] == "F":
		code[0] = "%05.0f" % (float(var[2]),)
		code[1] = "= "+str(int(var[3]))
		code[2] = "%02.0f" % (float(var[3]),)		
		code[3] = var[1]
	elif var[2] and var[3] and (var[4] in er) :
		code[0] = "%04.0f" % (float(var[2]),)
		code[1] = "= 0"
		code[2] = "000"		
		code[3] = "N"
	elif var[3] and var[4] and not var[1]:
		code[0] = "%04.0f" % (float(var[2]),)
		code[1] = "= "+str(var[3])+str(var[4])
		code[2] = str(var[3])+str(var[4])
		codeX = var[4]
		code[3] = "-"
	elif var[2] and var[3] and var[5] and not var[1]:
		code[0] = "%04.0f" % (float(var[2]),)
		code[1] = "= " + str(int(var[3]))
		code[2] = "%02.0f" % (float(var[3]),)
		codeX = "."+str(var[5])
		code[3] = "-"
	elif var[1] and var[2] and var[4] and var[5]:
		code[0] = "0000"
		code[1] = "= 0"
		code[2] = "000"
		code[3] = var[0]
	elif (var[2] and not var[1]):
		code[0] = "%04.0f" % (float(var[2]),)
		code[1] = "= " + str(int(var[3]))
		code[2] = "%03.0f" % (float(var[3]),)
		code[3] = "N"
	elif (var[2]):
		code[0] = "%04.0f" % (float(var[2]),)
		code[1] = "= " + str(int(var[3]))
		code[2] = "%03.0f" % (float(var[3]),)
		code[3] = var[1]
	else:
		code[0] = "0000"
		code[1] = "= 0"
		code[2] = "000"
		code[3] = "N"
	
	if ( code[3] == "NG" or code[3] == "ng" or code[3] == "N" ):
		code[3] = "N"
		at = "NG"
	elif (code[3] == "-"):
		code[3] = ""
		at = ""
	elif not code[3]:
		code[3] = -1
		at = "-1"
	else:
		code[3].upper()
		at = code[3]
 
	code[4] = str(code[3])+str(code[0])+"."+str(code[2])
	aan = int(code[0])
	aasn = str(code[2])+str(codeX)
	
	try:
	  ic2 = int(code[2])
	except:
	  ic2 = 0
	  	
	try:
	  fc2 = ("%02.0f" % (float(code[2]),))
	except:
	  fc2 = code[2]

	if (at == "NGL"):
		# Display NG Inventory No.
		code[5] = str(code[3])+str(code[0])+"."+str(code[2])
		code[6] = str(code[3])+"-"+str(code[0])+"-"+str(code[2])
		if (ic2):
		  code[14] = str(code[3])+str(code[0])+"."+str(code[2])
		else:
		  code[14] = str(code[3])+str(code[0])
	elif aasn:
		code[5] = str(at)+str(aan)+"."+str(aasn)
		code[6] = str(code[3])+"-"+str(aan)+"-"+str(aasn)
		if (ic2):
		  code[14] = str(at)+str(aan)+"."+str(aasn)
		else:
		  code[14] = str(at)+str(aan)
	elif aan:
		code[5] = str(at)+str(aan)
		code[6] = str(code[3])+"-"+str(aan)
		code[14] = str(at)+str(aan)
	else:
		code[5] = str(at)
		code[6] = str(at)
		code[14] = str(at)

	code[13] = code[5]
	
	# Not sure why this is done now, but wanted a simple F-number display for sample files
	if at == "F":
	  # Display NG Inventory No.
		code[5] = str(code[3])+str(code[0])+"."+str(code[2])
		code[6] = str(code[3])+"-"+str(code[0])+"-"+str(code[2])
		
	# NG painting related image filename base	

	code[7] = str(code[3])+"-"+str(code[0])+"-"+str(fc2)
	code[8] = ("%04.0f" % (float(code[0]),))+"-"+str(fc2)

	if aasn:
		code[9] = str(at)+str(code[0])+"."+str(code[2])
	else:
		code[9] = str(at)+str(code[0])

	if ic2 > 0:
			code[10] = code[7]
			code[11] = str(at)+str(code[0])+"_"+str(code[2])
	else:
			code[10] = str(code[3])+"-"+str(code[0])
			code[11] = str(at)+str(code[0])
	
	if not var[5]:
		var[5] = "No Match"
		
	if re.search('^[0-9]{6}$', var[5]):
		code[12] = str(code[7])+"-"+str(var[5])
	else:
		code[12] = False
	
	if not code[3] == -1:
		return code
	else:
		return False
